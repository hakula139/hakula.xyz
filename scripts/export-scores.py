#!/usr/bin/env python3
"""Export score data from Excel to CSV.

Reads xlsx files, evaluates cached formula values, detects formatting, and
outputs clean CSV files as Hugo page resources.

Usage:
    python3 scripts/export-scores.py
"""

from __future__ import annotations

import csv
import sys
from contextlib import closing
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any, TypedDict, cast

import openpyxl


if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.fills import PatternFill
    from openpyxl.styles.fonts import Font
    from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(__file__).resolve().parent.parent


class SourceConfig(TypedDict):
    xlsx: Path
    csv: Path
    headers: list[str]


SOURCES: dict[str, SourceConfig] = {
    'anime': {
        'xlsx': Path.home() / 'Hakula/Documents/Anime.xlsx',
        'csv': REPO_ROOT / 'content/posts/anime/impressions/scores.csv',
        'headers': [
            '作品名称',
            '年份',
            '剧本',
            '画面',
            '音乐',
            '总评',
            '评级',
            '个人差',
            'BGM',
            'Δ',
            '初次完成日期',
        ],
    },
    'avg': {
        'xlsx': Path.home() / 'Hakula/Documents/AVG.xlsx',
        'csv': REPO_ROOT / 'content/posts/avg/impressions/scores.csv',
        'headers': [
            '作品名称',
            '年份',
            '剧本',
            '系统',
            '画面',
            '音乐',
            '总评',
            '评级',
            '个人差',
            '分级',
            'EGS',
            'BGM',
            'Δ',
            '完成日期',
        ],
    },
}

# Columns that use 1 decimal place
DECIMAL_COLS = {'总评', '个人差', 'BGM', 'Δ'}
# Columns that use integer formatting
INT_COLS = {'年份', '剧本', '系统', '画面', '音乐', 'EGS'}
# Columns where gray text → prefix with ~ (approximate / insufficient data)
GRAY_PREFIX_COLS = {'EGS', 'BGM'}


def _parse_rgb(color_rgb: Any) -> tuple[int, int, int] | None:
    """Parse an openpyxl RGB value (e.g., "FF0000" or "FFFF0000") to (r, g, b)."""
    try:
        rgb = str(color_rgb)
    except (TypeError, ValueError):
        return None

    if len(rgb) >= 6:
        h = rgb[-6:]
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

    return None


def is_gray_font(font: Font) -> bool:
    """Check if font uses gray theme color (theme == 0 with negative tint)."""
    color = font.color
    return bool(color and color.type == 'theme' and color.theme == 0 and color.tint < 0)


def is_red_font(font: Font) -> bool:
    """Check if font color is explicit red (RGB, not theme)."""
    color = font.color
    if color and color.type == 'rgb':
        rgb = _parse_rgb(color.rgb)
        return rgb is not None and rgb[0] > 180 and rgb[1] < 100
    return False


def _parse_fill_rgb(fill: PatternFill) -> tuple[int, int, int] | None:
    """Parse solid fill foreground color to (r, g, b)."""
    if fill and cast(str, fill.fill_type) == 'solid' and fill.fgColor:
        return _parse_rgb(fill.fgColor.rgb)
    return None


def has_yellow_fill(fill: PatternFill) -> bool:
    """Check if cell has a solid yellow background fill."""
    rgb = _parse_fill_rgb(fill)
    return rgb is not None and rgb[0] > 180 and rgb[1] > 180 and rgb[2] < 100


def has_red_fill(fill: PatternFill) -> bool:
    """Check if cell has a solid red background fill."""
    rgb = _parse_fill_rgb(fill)
    return rgb is not None and rgb[0] > 180 and rgb[1] < 100 and rgb[2] < 100


def has_black_fill(fill: PatternFill) -> bool:
    """Check if cell has a solid black background fill."""
    rgb = _parse_fill_rgb(fill)
    return rgb is not None and rgb[0] < 50 and rgb[1] < 50 and rgb[2] < 50


def encode_age_rating(cell: Cell) -> str:
    """Encode age rating from cell formatting.

    Returns encoded string:

    - All, 15+, 17+                →  as-is    (level 0)
    - 18+ dark font                →  "18+"    (level 1)
    - 18+ red font                 →  "18+!"   (level 2)
    - 18+ red bold                 →  "18+!!"  (level 3)
    - 18+ red bold  + yellow fill  →  "18+G"   (level 4)
    - 18+ dark bold + red fill     →  "18+G!"  (level 5)
    - 18+ red bold  + black fill   →  "18+G!!" (level 6)
    """
    val = str(cell.value).strip() if cell.value else ''
    if not val:
        return val

    font = cast('Font', cell.font)
    fill = cast('PatternFill', cell.fill)
    red = is_red_font(font)
    bold = cast(bool, font.bold)

    if red and bold and has_black_fill(fill):
        return f'{val}G!!'
    if not red and bold and has_red_fill(fill):
        return f'{val}G!'
    if red and bold and has_yellow_fill(fill):
        return f'{val}G'
    if red and bold:
        return f'{val}!!'
    if red:
        return f'{val}!'
    return val


def format_decimal(
    val: int | float,
    *,
    places: int = 1,
    sign: bool = False,
) -> str:
    """Round a float using half-up rounding and format."""
    d = Decimal(str(val))
    q = Decimal('0.' + '0' * places)
    result = str(d.quantize(q, rounding=ROUND_HALF_UP))
    if sign and val > 0:
        return f'+{result}'
    return result


def format_value(val: Any, header: str) -> str:
    """Format a cell value according to column type."""
    if val is None:
        return ''

    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')

    if header == 'Δ':
        if isinstance(val, (int, float)):
            return format_decimal(val, sign=True)
        return str(val)

    if header in DECIMAL_COLS:
        if isinstance(val, (int, float)):
            return format_decimal(val)
        return str(val)

    if header in INT_COLS:
        if isinstance(val, (int, float)):
            return format_decimal(val, places=0)
        return str(val)

    return str(val).strip()


def find_header_columns(
    sheet: Worksheet, expected_headers: list[str]
) -> dict[str, int]:
    """Find column indices by matching header names in the first row.

    Returns a dict mapping header name → column index (1-based).
    """
    header_map: dict[str, int] = {}
    for col_idx in range(1, sheet.max_column + 1):
        if cell_val := sheet.cell(row=1, column=col_idx).value:
            cell_val = str(cell_val).strip()
            if cell_val in expected_headers:
                header_map[cell_val] = col_idx
    return header_map


def export_source(name: str, config: SourceConfig) -> bool:
    """Export a single xlsx source to CSV."""
    xlsx_path = config['xlsx']
    csv_path = config['csv']
    headers = config['headers']

    if not xlsx_path.exists():
        print(f'Skipping {name}: {xlsx_path} not found', file=sys.stderr)
        return False

    print(f'Reading {xlsx_path}')

    with (
        closing(openpyxl.load_workbook(xlsx_path, data_only=True)) as wb,
        closing(openpyxl.load_workbook(xlsx_path, data_only=False)) as wb_fmt,
    ):
        sheet = wb.active
        sheet_fmt = wb_fmt.active
        if sheet is None or sheet_fmt is None:
            print(f'Skipping {name}: no active sheet in {xlsx_path}', file=sys.stderr)
            return False

        # Find column positions from headers
        col_map = find_header_columns(sheet, headers)
        missing = [h for h in headers if h not in col_map]
        if missing:
            print(f'Warning: columns not found: {missing}', file=sys.stderr)

        # Find the date column for this source
        date_headers = [h for h in headers if h.endswith('完成日期')]
        date_col = col_map.get(date_headers[0]) if date_headers else None

        rows_out: list[list[str]] = []
        for row_idx in range(2, sheet.max_row + 1):
            # Skip rows without a completion date
            if date_col:
                date_val = sheet.cell(row=row_idx, column=date_col).value
                if not date_val:
                    continue

            row_data: list[str] = []
            for header in headers:
                col_idx = col_map.get(header)
                if col_idx is None:
                    row_data.append('')
                    continue

                cell_fmt = cast('Cell', sheet_fmt.cell(row=row_idx, column=col_idx))

                if header == '分级':
                    row_data.append(encode_age_rating(cell_fmt))
                    continue

                val = sheet.cell(row=row_idx, column=col_idx).value
                formatted = format_value(val, header)

                # Gray font indicates approximate or insufficient data;
                # prefix the value with ~ so the frontend can style it.
                if (
                    header in GRAY_PREFIX_COLS
                    and formatted
                    and is_gray_font(cast('Font', cell_fmt.font))
                ):
                    formatted = f'~{formatted}'

                row_data.append(formatted)

            rows_out.append(row_data)

    # Write CSV
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows_out)

    print(f'Wrote {len(rows_out)} rows to {csv_path}')
    return True


def main() -> None:
    print('Exporting score data from Excel to CSV...')
    for name, config in SOURCES.items():
        print(f'\n[{name}]')
        export_source(name, config)
    print('\nDone.')


if __name__ == '__main__':
    main()
